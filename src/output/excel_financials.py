"""Build IS + Valuation workbook matching SEC statement-of-operations layout.

The IS sheet renders the company's XBRL-reported line items in SEC order,
with forecast columns (2026E onwards) populated from the DCF.
"""
from __future__ import annotations
from pathlib import Path
from typing import Dict, List, Optional

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from ..extract.income_statement import IncomeStatement, ISLine, LABEL_OVERRIDES, _display_label
from ..model.dcf import DCFResult


# ---------- Styling ----------

HEADER_FILL = PatternFill("solid", fgColor="1F3864")
HEADER_FONT = Font(name="Calibri", size=11, color="FFFFFF", bold=True)
COMPANY_FONT = Font(name="Calibri", size=12, color="1F3864", bold=True)
FORECAST_FILL = PatternFill("solid", fgColor="DDEBF7")  # light blue for forecast columns
SECTION_FONT = Font(name="Calibri", size=10, bold=True, color="1F3864")
LABEL_FONT = Font(name="Calibri", size=10)
BOLD_LABEL_FONT = Font(name="Calibri", size=10, bold=True)
INPUT_FONT = Font(name="Calibri", size=10, color="0070C0")
FORECAST_FONT = Font(name="Calibri", size=10, italic=True, color="0070C0")
TOTAL_FONT = Font(name="Calibri", size=10, bold=True)
ITALIC_FONT = Font(name="Calibri", size=10, italic=True, color="595959")

THIN = Side(border_style="thin", color="BFBFBF")
TOP_BORDER = Border(top=Side(border_style="thin", color="000000"))
DOUBLE_TOP = Border(top=Side(border_style="double", color="000000"))

RIGHT_ALIGN = Alignment(horizontal="right")
CENTER_ALIGN = Alignment(horizontal="center")
LEFT_ALIGN = Alignment(horizontal="left")

FMT_NUM = "#,##0;(#,##0);\"-\""
FMT_PCT = "0%"
FMT_PCT1 = "0.0%"
FMT_USD = "$#,##0.00"
FMT_SHARES = "#,##0"


# ---------- Helpers ----------

def _all_period_labels(stmt: IncomeStatement, dcf: "DCFResult") -> List[str]:
    """Historical periods + base-case forecast year labels."""
    labels = [p.label for p in stmt.periods]
    labels += [f"{p.year}E" for p in dcf.base.projections]
    return labels


def _forecast_labels_set(dcf: "DCFResult") -> set:
    return {f"{p.year}E" for p in dcf.base.projections}


def _build_forecast_values(stmt: IncomeStatement, dcf: "DCFResult") -> Dict[str, Dict[str, float]]:
    """Project values for forecast columns in the IS (base case only).

    Only the skeleton totals are projected:
      revenue_total → revenue
      opex_total    → revenue - EBIT
      operating_income → EBIT
      pre_tax       → EBIT (assumes no non-op for forecast)
      tax           → tax line
      net_income    → NOPAT
    """
    result: Dict[str, Dict[str, float]] = {
        "revenue_total": {},
        "opex_total": {},
        "operating_income": {},
        "pre_tax": {},
        "tax": {},
        "net_income": {},
    }
    for p in dcf.base.projections:
        label = f"{p.year}E"
        result["revenue_total"][label] = p.revenue
        result["opex_total"][label] = p.revenue - p.ebit
        result["operating_income"][label] = p.ebit
        result["pre_tax"][label] = p.ebit
        result["tax"][label] = p.tax
        result["net_income"][label] = p.nopat
    return result


def _write_header(ws, period_labels: List[str], forecast_labels: set, title: str, ticker: str):
    """Top header rows with company, ticker, period labels. Forecast columns shaded."""
    ws.cell(row=1, column=2, value=title).font = COMPANY_FONT
    ws.cell(row=2, column=2, value=ticker).font = COMPANY_FONT
    ws.cell(row=3, column=2, value="in US$ '000").font = ITALIC_FONT

    for col_idx, label in enumerate(period_labels, start=3):
        cell = ws.cell(row=4, column=col_idx, value=label)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER_ALIGN
        if label in forecast_labels:
            # add extra light-blue shading on the header for forecast columns (already navy, skip)
            pass


def _write_line_values(ws, row: int, label: str, values: Dict[str, Optional[float]],
                       period_labels: List[str], forecast_labels: set,
                       *, bold=False, italic=False, fmt=FMT_NUM, divide_by=1000.0,
                       forecast_values: Optional[Dict[str, float]] = None,
                       forecast_formulas: Optional[Dict[str, str]] = None):
    """Write one line of the IS with historical values and forecast cells.

    forecast_formulas (optional): {period_label: "=..."} — if provided, the
    forecast cell is a live Excel formula instead of a hardcoded value.
    """
    label_cell = ws.cell(row=row, column=2, value=label)
    label_cell.font = BOLD_LABEL_FONT if bold else (ITALIC_FONT if italic else LABEL_FONT)
    label_cell.alignment = LEFT_ALIGN

    for col_idx, p_label in enumerate(period_labels, start=3):
        is_forecast = p_label in forecast_labels
        cell_value = None
        is_formula = False

        if is_forecast and forecast_formulas and p_label in forecast_formulas:
            cell_value = forecast_formulas[p_label]
            is_formula = True
        elif is_forecast and forecast_values and p_label in forecast_values:
            v = forecast_values[p_label]
            cell_value = (v / divide_by) if fmt in (FMT_NUM, FMT_SHARES) else v
        elif not is_forecast:
            v = values.get(p_label)
            if v is not None:
                cell_value = (v / divide_by) if fmt in (FMT_NUM, FMT_SHARES) else v

        if cell_value is None:
            if is_forecast:
                ws.cell(row=row, column=col_idx).fill = FORECAST_FILL
            continue

        cell = ws.cell(row=row, column=col_idx, value=cell_value)
        cell.number_format = fmt
        cell.alignment = RIGHT_ALIGN
        if is_forecast:
            cell.font = TOTAL_FONT if bold else FORECAST_FONT
            cell.fill = FORECAST_FILL
        else:
            cell.font = TOTAL_FONT if bold else LABEL_FONT


def _write_section_heading(ws, row: int, text: str, n_cols: int):
    c = ws.cell(row=row, column=2, value=text)
    c.font = SECTION_FONT


def _write_margin_formula_row(ws, row: int, label: str, numerator_row: int,
                               denominator_row: int, period_labels: List[str]):
    """Write a row where each column is = numerator_row / denominator_row (as formula)."""
    c = ws.cell(row=row, column=2, value=label)
    c.font = ITALIC_FONT; c.alignment = LEFT_ALIGN
    for col_idx, _ in enumerate(period_labels, start=3):
        L = get_column_letter(col_idx)
        formula = f"=IFERROR({L}{numerator_row}/{L}{denominator_row},\"\")"
        cell = ws.cell(row=row, column=col_idx, value=formula)
        cell.font = ITALIC_FONT
        cell.number_format = FMT_PCT
        cell.alignment = RIGHT_ALIGN


def _write_yoy_formula_row(ws, row: int, label: str, source_row: int,
                            period_labels: List[str], periods_by_label: dict):
    """Write YoY growth formulas referencing the source IS row.

    For annual periods: =curr / prior_year_same_col_via_label - 1
    For quarters: =curr / same_quarter_prior_year - 1
    Implementation: we precompute which IS column holds the prior-year period
    for each column, and emit an Excel formula only when there's a matching prior column.
    """
    label_cell = ws.cell(row=row, column=2, value=label)
    label_cell.font = ITALIC_FONT; label_cell.alignment = LEFT_ALIGN

    # Map period_label → column letter in this IS sheet
    label_to_col = {lbl: get_column_letter(3 + i) for i, lbl in enumerate(period_labels)}

    for col_idx, p_label in enumerate(period_labels, start=3):
        # Determine the "prior period" label
        if p_label in periods_by_label:
            p = periods_by_label[p_label]
            if p.is_annual:
                prev_label = str(p.fiscal_year - 1)
            else:
                prev_label = f"{p.quarter}Q{str(p.fiscal_year - 1)[-2:]}"
        else:
            # Forecast column like "2026E"
            try:
                year_str = p_label.replace("E", "")
                prev_year = int(year_str) - 1
                # prior year could be an annual column (e.g. "2025") or forecast "2025E"
                prev_label = str(prev_year)
                if prev_label not in label_to_col:
                    prev_label = f"{prev_year}E"
            except ValueError:
                continue

        if prev_label not in label_to_col:
            continue

        curr_col = get_column_letter(col_idx)
        prev_col = label_to_col[prev_label]
        formula = f"=IFERROR({curr_col}{source_row}/{prev_col}{source_row}-1,\"\")"
        cell = ws.cell(row=row, column=col_idx, value=formula)
        cell.font = ITALIC_FONT
        cell.number_format = FMT_PCT
        cell.alignment = RIGHT_ALIGN


def _yoy_growth(values: Dict[str, float], period_labels: List[str],
                periods_by_label: Dict[str, "Period"]) -> Dict[str, float]:
    """Compute YoY growth for each period (same-quarter or year-over-year)."""
    result = {}
    for label in period_labels:
        v = values.get(label)
        if v is None:
            continue
        p = periods_by_label.get(label)
        if p is None:
            continue
        if p.is_annual:
            prev_label = str(p.fiscal_year - 1)
        else:
            prev_label = f"{p.quarter}Q{str(p.fiscal_year - 1)[-2:]}"
        prev = values.get(prev_label)
        if prev and prev != 0:
            result[label] = (v - prev) / abs(prev)
    return result


# ---------- IS sheet ----------

def _write_is_sheet(ws, stmt: IncomeStatement, dcf: DCFResult):
    period_labels = _all_period_labels(stmt, dcf)
    forecast_labels = _forecast_labels_set(dcf)
    periods_by_label = {p.label: p for p in stmt.periods}

    # Column widths
    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 54
    for c in range(3, 3 + len(period_labels)):
        ws.column_dimensions[get_column_letter(c)].width = 12

    _write_header(ws, period_labels, forecast_labels, stmt.title, stmt.ticker)

    # Build forecast skeleton values from DCF (used for margin / YoY derivations)
    forecast_by_section = _build_forecast_values(stmt, dcf)

    # Build forecast FORMULAS that link into the Valuation sheet.
    # Valuation sheet projection table rows: revenue=56, EBIT=58, Tax=59, NOPAT=60
    # (column D is Y1, E is Y2, ..., L is Y10). Values there are in $M,
    # IS shows in $'000 (k), so multiply by 1000.
    VAL_REV_ROW, VAL_EBIT_ROW, VAL_TAX_ROW, VAL_NOPAT_ROW = 56, 58, 59, 60
    forecast_formula_by_section: Dict[str, Dict[str, str]] = {
        "revenue_total":    {},
        "opex_total":       {},
        "operating_income": {},
        "pre_tax":          {},
        "tax":              {},
        "net_income":       {},
    }
    # Forecast year i (1..total) → Valuation sheet column letter.
    # Year 1 = column D on Valuation sheet (col index 4), year 2 = E (5), etc.
    for i, p in enumerate(dcf.base.projections, start=1):
        val_col = get_column_letter(3 + i)   # column D for year 1 (index 4)
        label = f"{p.year}E"
        forecast_formula_by_section["revenue_total"][label]    = f"=Valuation!{val_col}{VAL_REV_ROW}*1000"
        forecast_formula_by_section["operating_income"][label] = f"=Valuation!{val_col}{VAL_EBIT_ROW}*1000"
        # Total opex = revenue - EBIT
        forecast_formula_by_section["opex_total"][label]       = f"=(Valuation!{val_col}{VAL_REV_ROW}-Valuation!{val_col}{VAL_EBIT_ROW})*1000"
        # Pre-tax income ≈ EBIT for forecast (no non-op modeled)
        forecast_formula_by_section["pre_tax"][label]          = f"=Valuation!{val_col}{VAL_EBIT_ROW}*1000"
        # Tax = -Valuation tax row (tax stored negative there, we want positive on IS)
        forecast_formula_by_section["tax"][label]              = f"=-Valuation!{val_col}{VAL_TAX_ROW}*1000"
        # Net income ≈ NOPAT for forecast (proxy since we don't model non-op)
        forecast_formula_by_section["net_income"][label]       = f"=Valuation!{val_col}{VAL_NOPAT_ROW}*1000"

    # Render sections in order, grouping lines by section
    row = 6
    sections_order = ["revenue", "revenue_total", "cost_of_revenue", "gross_profit",
                      "opex", "opex_adjustment", "opex_total", "operating_income",
                      "nonop", "pre_tax", "tax", "net_income", "eps", "shares"]

    lines_by_section: Dict[str, List[ISLine]] = {s: [] for s in sections_order}
    for line in stmt.lines:
        if line.section in lines_by_section:
            lines_by_section[line.section].append(line)

    # Track row numbers of key lines so we can emit formulas that reference them.
    section_first_row: Dict[str, int] = {}

    section_headings = {
        "revenue": "Revenues",
        "opex": "Operating expenses",
        "nonop": "Non-operating income (expense)",
        "eps": "Per-share data",
    }
    added_headings = set()

    for section in sections_order:
        lines = lines_by_section[section]
        # Top-of-section heading for certain groups
        if section in section_headings and lines and section not in added_headings:
            _write_section_heading(ws, row, section_headings[section], len(period_labels))
            added_headings.add(section)
            row += 1

        for line in lines:
            fv = forecast_by_section.get(section) if line.is_total or section in (
                "revenue_total", "opex_total", "operating_income", "pre_tax", "tax", "net_income"
            ) else None
            # Only link via formula for the primary (first) line of these skeleton
            # sections; sub-lines (individual opex items, etc.) don't have a
            # corresponding row on the Valuation sheet.
            ff = forecast_formula_by_section.get(section) if line.is_total or section in (
                "revenue_total", "opex_total", "operating_income", "pre_tax", "tax", "net_income"
            ) else None

            fmt, divisor = FMT_NUM, 1000.0
            if section == "eps":
                fmt, divisor = FMT_USD, 1.0
            elif section == "shares":
                fmt, divisor = FMT_SHARES, 1000.0

            _write_line_values(
                ws, row, line.label, line.values, period_labels, forecast_labels,
                bold=line.is_total, fmt=fmt, divide_by=divisor,
                forecast_values=fv, forecast_formulas=ff,
            )
            if line.is_total:
                for c in range(2, 3 + len(period_labels)):
                    ws.cell(row=row, column=c).border = TOP_BORDER
            # Track first row per section (primary line of each)
            if section not in section_first_row:
                section_first_row[section] = row
            row += 1

        # After Total revenue, insert YoY row (formula-driven)
        if section == "revenue_total" and lines:
            rev_row = section_first_row.get("revenue_total")
            _write_yoy_formula_row(ws, row, "Δ in revenue (YoY)",
                                   source_row=rev_row, period_labels=period_labels,
                                   periods_by_label=periods_by_label)
            row += 2

        # After Total costs and expenses, insert YoY row
        if section == "opex_total" and lines:
            opex_row = section_first_row.get("opex_total")
            _write_yoy_formula_row(ws, row, "Δ in operating expenses (YoY)",
                                   source_row=opex_row, period_labels=period_labels,
                                   periods_by_label=periods_by_label)
            row += 2

        # After Operating income, insert Operating margin (= operating income / revenue)
        if section == "operating_income" and lines:
            op_row = section_first_row.get("operating_income")
            rev_row = section_first_row.get("revenue_total")
            if op_row and rev_row:
                _write_margin_formula_row(ws, row, "Margin", numerator_row=op_row,
                                          denominator_row=rev_row,
                                          period_labels=period_labels)
                row += 2

        # After Net income, insert Net margin (= net income / revenue)
        if section == "net_income" and lines:
            ni_row = section_first_row.get("net_income")
            rev_row = section_first_row.get("revenue_total")
            if ni_row and rev_row:
                _write_margin_formula_row(ws, row, "Margin", numerator_row=ni_row,
                                          denominator_row=rev_row,
                                          period_labels=period_labels)
                row += 2

    # Company Financials block
    row += 1
    _write_section_heading(ws, row, "Company Financials", len(period_labels))
    row += 1

    cf_items = [
        ("balance_sheet", "cash_and_equivalents", "Cash and cash equivalents"),
        ("balance_sheet", "long_term_debt", "Long-term debt"),
        ("balance_sheet", "short_term_debt", "Short-term debt"),
        ("cash_flow", "d_and_a", "Depreciation and amortization"),
        ("cash_flow", "capex", "CapEx"),
        ("cash_flow", "operating_cash_flow", "Cash from operations"),
    ]
    for bucket_name, key, label in cf_items:
        bucket = getattr(stmt, bucket_name)
        if key in bucket:
            _write_line_values(ws, row, label, bucket[key], period_labels,
                               forecast_labels, fmt=FMT_NUM)
            row += 1

    # EBITDA = Operating income + D&A (formula linking to OI row and D&A row above)
    op_row = section_first_row.get("operating_income")
    da_row_written: Optional[int] = None
    # Walk back up to find the "Depreciation and amortization" row we just wrote
    # in the Company Financials block.
    for r in range(row - 1, max(row - 10, 1), -1):
        label = ws.cell(row=r, column=2).value
        if label and "Depreciation" in str(label) and "amortization" in str(label).lower():
            da_row_written = r
            break

    if op_row and da_row_written:
        ws.cell(row=row, column=2, value="EBITDA").font = TOTAL_FONT
        for col_idx, p_label in enumerate(period_labels, start=3):
            L = get_column_letter(col_idx)
            # Only emit formula if both sources have a value; otherwise leave blank
            op_cell_val = ws.cell(row=op_row, column=col_idx).value
            da_cell_val = ws.cell(row=da_row_written, column=col_idx).value
            if op_cell_val is None and da_cell_val is None:
                continue
            # Use IFERROR to gracefully handle missing periods
            formula = f"=IFERROR({L}{op_row}+{L}{da_row_written},{L}{op_row})"
            cell = ws.cell(row=row, column=col_idx, value=formula)
            cell.font = TOTAL_FONT
            cell.number_format = FMT_NUM
            cell.alignment = RIGHT_ALIGN
            if p_label in forecast_labels:
                cell.fill = FORECAST_FILL
        for c in range(2, 3 + len(period_labels)):
            ws.cell(row=row, column=c).border = TOP_BORDER
        row += 1

    ws.freeze_panes = "C5"


# ---------- Valuation sheet (bank-style, FORMULA-LIVE) ----------
#
# Every derived cell on the Valuation sheet chains through Excel formulas.
# Inputs live at the top in light-blue; change them and the whole model
# recomputes. The scenario selector cell (C40) toggles Bear/Base/Bull by
# pointing the projection formulas at the appropriate adjustment column
# via INDEX/MATCH.
#
# Row map (key anchors referenced in formulas):
#   C5  rf,  C6  ERP,  C7  β,  C8  tax
#   C11 trailing CAGR, C12 current EBIT margin
#   C13 D&A%, C14 CapEx%, C15 SBC%, C16 NWC%
#   C19 explicit yrs, C20 fade yrs, C21 terminal g,
#   C22 TV weight Gordon, C23 TV weight Exit, C24 peer EV/EBITDA
#   C27 market cap, C28 debt, C29 cash, C30 shares, C31 base revenue
#   D34:F34 scenario labels, D35:F35 growth adj, D36:F36 margin adj, D37:F37 WACC adj
#   C40 scenario selector (type "Base" / "Bull" / "Bear")
#   C43 credit spread, C44 Ke, C45 Kd pretax, C46 Kd aftertax,
#   C47 E/V, C48 D/V, C49 WACC(selected)
#   Row 52-67: projection table columns C (Y0) ... M (Y10)
#   Row 70+: terminal value, equity bridge, scenario summary

def _write_valuation_sheet(ws, stmt: IncomeStatement, dcf: "DCFResult",
                            current_share_price: Optional[float],
                            current_market_cap: Optional[float],
                            football_field_png: Optional[Path] = None):
    from ..model.dcf import DCFResult, DCFScenarioResult
    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 38
    for col in range(3, 16):
        ws.column_dimensions[get_column_letter(col)].width = 12

    from ..model.dcf import DCFResult, DCFScenarioResult
    inputs_fill = PatternFill("solid", fgColor="DDEBF7")
    calc_fill = PatternFill("solid", fgColor="FFFFFF")
    section_fill = PatternFill("solid", fgColor="F2F2F2")

    def _input(r, label, val, fmt):
        ws.cell(row=r, column=2, value=label).font = LABEL_FONT
        c = ws.cell(row=r, column=3, value=val)
        c.font = INPUT_FONT
        c.fill = inputs_fill
        c.number_format = fmt
        c.alignment = RIGHT_ALIGN
        return c

    def _calc(r, label, formula, fmt, bold=False):
        ws.cell(row=r, column=2, value=label).font = TOTAL_FONT if bold else LABEL_FONT
        c = ws.cell(row=r, column=3, value=formula)
        c.font = TOTAL_FONT if bold else LABEL_FONT
        c.fill = calc_fill
        c.number_format = fmt
        c.alignment = RIGHT_ALIGN
        return c

    def _section(r, text, col=2):
        c = ws.cell(row=r, column=col, value=text)
        c.font = SECTION_FONT
        c.fill = section_fill

    # ===== Title =====
    ws.cell(row=1, column=2, value=f"{stmt.title}  —  DCF Valuation").font = Font(
        name="Calibri", size=14, color="1F3864", bold=True)
    ws.cell(row=2, column=2,
            value=f"Ticker: {stmt.ticker}   |   Base FY: {dcf.base_fy}   |   "
                  f"Methodology: {dcf.assumptions.explicit_years}Y explicit + "
                  f"{dcf.assumptions.fade_years}Y fade + dual terminal value").font = ITALIC_FONT
    ws.cell(row=3, column=2,
            value="Blue cells = inputs (editable). White cells = formulas. "
                  "Change inputs → model recomputes.").font = ITALIC_FONT

    # ============================================================
    # BLOCK 1 — Market Inputs (rows 4-8)
    # ============================================================
    _section(4, "Market Inputs (CAPM)")
    w = dcf.wacc_base
    _input(5, "Risk-free rate (10Y Treasury)", w.inputs.risk_free_rate, FMT_PCT1)
    _input(6, "Equity risk premium",            w.inputs.equity_risk_premium, FMT_PCT1)
    _input(7, "Beta (levered)",                 w.inputs.beta, "0.00")
    _input(8, "Tax rate",                       dcf.assumptions.tax_rate, FMT_PCT1)

    # ============================================================
    # BLOCK 2 — Operating Assumptions (rows 10-16)
    # Column C = trailing ratio (used Y1-Y2). Column D = terminal ratio (year 7+
    # steady state). CapEx and SBC fade linearly from trailing → terminal
    # over years 3-7 in the projection.
    # ============================================================
    _section(10, "Operating Assumptions")
    # Column headers for trailing vs terminal
    ws.cell(row=10, column=3, value="Trailing").font = ITALIC_FONT
    ws.cell(row=10, column=3).alignment = RIGHT_ALIGN
    ws.cell(row=10, column=4, value="Terminal (Y7+)").font = ITALIC_FONT
    ws.cell(row=10, column=4).alignment = RIGHT_ALIGN

    _input(11, "Trailing 3Y revenue CAGR",     dcf.trailing_revenue_cagr, FMT_PCT1)
    _input(12, "Current EBIT margin",          dcf.current_ebit_margin, FMT_PCT1)
    _input(13, "D&A as % of revenue",           dcf.da_pct_revenue, FMT_PCT1)
    _input(14, "CapEx as % of revenue",         dcf.capex_pct_revenue, FMT_PCT1)
    # Terminal CapEx ratio (column D, same row)
    c = ws.cell(row=14, column=4, value=dcf.assumptions.terminal_capex_pct_revenue)
    c.font = INPUT_FONT; c.fill = inputs_fill
    c.number_format = FMT_PCT1; c.alignment = RIGHT_ALIGN

    _input(15, "SBC as % of revenue",           dcf.sbc_pct_revenue, FMT_PCT1)
    # Terminal SBC ratio
    c = ws.cell(row=15, column=4, value=dcf.assumptions.terminal_sbc_pct_revenue)
    c.font = INPUT_FONT; c.fill = inputs_fill
    c.number_format = FMT_PCT1; c.alignment = RIGHT_ALIGN

    _input(16, "ΔNWC as % of ΔRevenue",         dcf.nwc_pct_delta_revenue, FMT_PCT1)

    # ============================================================
    # BLOCK 3 — DCF Parameters (rows 18-24)
    # ============================================================
    _section(18, "DCF Parameters")
    total_years = dcf.assumptions.explicit_years + dcf.assumptions.fade_years
    _input(19, "Explicit years",                dcf.assumptions.explicit_years, "0")
    _input(20, "Fade years",                    dcf.assumptions.fade_years, "0")
    _input(21, "Terminal growth rate",          dcf.assumptions.terminal_growth, FMT_PCT1)
    _input(22, "TV weight — Gordon Growth",     dcf.assumptions.tv_weight_gordon, "0.00")
    _input(23, "TV weight — Exit Multiple",     dcf.assumptions.tv_weight_exit, "0.00")
    _input(24, "Peer median EV/EBITDA (exit)",  dcf.peer_median_ev_ebitda or 0.0, "0.0\"x\"")

    # ============================================================
    # BLOCK 4 — Capital Structure (rows 26-31)  [in $M]
    # ============================================================
    _section(26, "Capital Structure ($M, in millions)")
    _input(27, "Market cap",             (current_market_cap or 0) / 1e6, FMT_NUM)
    _input(28, "Total debt",             dcf.debt / 1e6, FMT_NUM)
    _input(29, "Cash & equivalents",     dcf.cash / 1e6, FMT_NUM)
    _input(30, "Shares outstanding (mm)", dcf.shares_outstanding / 1e6, FMT_SHARES)
    _input(31, "Base-year revenue",      dcf.base_revenue / 1e6, FMT_NUM)

    # ============================================================
    # Analyst Consensus block (columns E-G, rows 10-18)
    # These are the CELLS that drive the projection — edit to override.
    # Anchored cell references (used by projection formulas):
    #   G13 = Consensus Y1 growth   (blank → fallback to trailing CAGR C11)
    #   G14 = Consensus Y2 growth   (blank → fallback)
    #   G15 = Target EBIT margin    (blank → fallback to current margin C12)
    # ============================================================
    _section(10, "Analyst Consensus (sell-side) — drives forecast", col=5)
    rec = dcf.analyst_rec_key or "n/a"
    n = dcf.analyst_count or 0

    # Textual rows
    ws.cell(row=11, column=5, value="Analyst rating").font = LABEL_FONT
    ws.cell(row=11, column=7, value=rec.replace("_", " ").title() if rec != "n/a" else "n/a").font = LABEL_FONT
    ws.cell(row=12, column=5, value="Number of analysts").font = LABEL_FONT
    ws.cell(row=12, column=7, value=n).number_format = "0"
    ws.cell(row=12, column=7).font = LABEL_FONT

    # Driver rows (G13-G16 = input cells that drive the projection formulas)
    for r_idx, (label, val, fmt) in enumerate([
        ("Revenue growth — Y1 (consensus)",        dcf.analyst_growth_y1,       FMT_PCT1),
        ("Revenue growth — Y2 (consensus)",        dcf.analyst_growth_y2,       FMT_PCT1),
        ("Target EBIT margin Y1 (from EPS Y1)",    dcf.target_ebit_margin_y1,   FMT_PCT1),
        ("Target EBIT margin Y2 (from EPS Y2)",    dcf.target_ebit_margin_y2,   FMT_PCT1),
    ], start=13):
        ws.cell(row=r_idx, column=5, value=label).font = LABEL_FONT
        if val is not None:
            c = ws.cell(row=r_idx, column=7, value=val)
            c.font = INPUT_FONT
            c.fill = inputs_fill
            c.number_format = fmt
            c.alignment = RIGHT_ALIGN

    # Price targets (informational only)
    for r_idx, (label, val, fmt) in enumerate([
        ("Price target — low",  dcf.analyst_target_low,  FMT_USD),
        ("Price target — mean", dcf.analyst_target_mean, FMT_USD),
        ("Price target — high", dcf.analyst_target_high, FMT_USD),
    ], start=17):
        ws.cell(row=r_idx, column=5, value=label).font = LABEL_FONT
        if val is not None:
            c = ws.cell(row=r_idx, column=7, value=val)
            c.font = LABEL_FONT
            c.number_format = fmt
            c.alignment = RIGHT_ALIGN

    # ============================================================
    # BLOCK 5 — Scenario Adjustments + Selector (rows 33-40)
    # ============================================================
    _section(33, "Scenario Adjustments")
    # Column headers D=Bear E=Base F=Bull
    for i, name in enumerate(["Bear", "Base", "Bull"]):
        c = ws.cell(row=34, column=4 + i, value=name)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = CENTER_ALIGN
    # Row 35: revenue growth adj
    ws.cell(row=35, column=2, value="Revenue growth starting adj (vs CAGR)").font = LABEL_FONT
    for i, val in enumerate([-0.03, 0.00, 0.03]):
        c = ws.cell(row=35, column=4 + i, value=val)
        c.font = INPUT_FONT; c.fill = inputs_fill
        c.number_format = FMT_PCT1; c.alignment = RIGHT_ALIGN
    # Row 36: EBIT margin premium
    ws.cell(row=36, column=2, value="EBIT margin premium (vs current)").font = LABEL_FONT
    for i, val in enumerate([dcf.assumptions.ebit_margin_premium_bear,
                              dcf.assumptions.ebit_margin_premium_base,
                              dcf.assumptions.ebit_margin_premium_bull]):
        c = ws.cell(row=36, column=4 + i, value=val)
        c.font = INPUT_FONT; c.fill = inputs_fill
        c.number_format = FMT_PCT1; c.alignment = RIGHT_ALIGN
    # Row 37: WACC adjustment
    ws.cell(row=37, column=2, value="WACC adjustment").font = LABEL_FONT
    for i, val in enumerate([dcf.assumptions.wacc_adjustment_bear,
                              0.0,
                              dcf.assumptions.wacc_adjustment_bull]):
        c = ws.cell(row=37, column=4 + i, value=val)
        c.font = INPUT_FONT; c.fill = inputs_fill
        c.number_format = FMT_PCT1; c.alignment = RIGHT_ALIGN

    # Row 39-40: Scenario selector
    ws.cell(row=39, column=2, value="Active scenario (type Bear / Base / Bull)").font = BOLD_LABEL_FONT
    sel = ws.cell(row=40, column=3, value="Base")
    sel.font = INPUT_FONT; sel.fill = inputs_fill; sel.alignment = Alignment(horizontal="center")
    sel.number_format = "@"
    # Store into a defined name for readability
    ws.cell(row=40, column=2, value="Scenario selector →").font = BOLD_LABEL_FONT
    ws.cell(row=40, column=2).alignment = Alignment(horizontal="right")

    # Helper: INDEX/MATCH formula string to fetch scenario adjustment from row N
    # MATCH is against D34:F34 (Bear/Base/Bull headers).
    def scen_lookup(row_idx):
        return f"INDEX($D${row_idx}:$F${row_idx},MATCH($C$40,$D$34:$F$34,0))"

    # ============================================================
    # BLOCK 6 — WACC Derivation (rows 42-49)
    # ============================================================
    _section(42, "WACC Derivation (CAPM)")
    _input(43, "Credit spread (leverage-bucketed)", w.credit_spread, FMT_PCT1)
    _calc(44, "Cost of equity  Ke = rf + β × ERP",   "=C5+C7*C6", FMT_PCT1)
    _calc(45, "Pre-tax cost of debt  Kd = rf + spread", "=C5+C43", FMT_PCT1)
    _calc(46, "After-tax cost of debt  Kd × (1−t)",   "=C45*(1-C8)", FMT_PCT1)
    _calc(47, "Equity weight  E/V",                    "=C27/(C27+C28)", FMT_PCT1)
    _calc(48, "Debt weight  D/V",                      "=C28/(C27+C28)", FMT_PCT1)
    # WACC uses scenario-selected WACC adjustment
    _calc(49, "WACC (selected scenario)",
          f"=C47*C44+C48*C46+{scen_lookup(37)}", FMT_PCT1, bold=True)

    # ============================================================
    # BLOCK 7 — Projection Table (rows 52-68)
    # Columns: C=Y0 (base year), D=Y1, ..., M=Y10
    # ============================================================
    _section(52, f"Base Case Projections ($M, Y0 actual, Y1–Y{total_years} forecast)")
    year_header_row = 53
    year_index_row = 54
    growth_row = 55
    revenue_row = 56
    margin_row = 57
    ebit_row = 58
    tax_row = 59
    nopat_row = 60
    da_row = 61
    capex_row = 62
    sbc_row = 63
    nwc_row = 64
    fcff_row = 65
    disc_row = 66
    pv_row = 67
    ebitda_row = 68

    # Column C is Y0; D through (C+total_years) are Y1..Yn
    last_col_idx = 3 + total_years        # column M when total=10
    last_col_letter = get_column_letter(last_col_idx)

    # Header row: year
    c = ws.cell(row=year_header_row, column=3, value=dcf.base_fy)
    c.font = HEADER_FONT; c.fill = HEADER_FILL; c.alignment = CENTER_ALIGN
    for i in range(total_years):
        col = 4 + i
        c = ws.cell(row=year_header_row, column=col, value=dcf.base_fy + 1 + i)
        c.font = HEADER_FONT; c.fill = HEADER_FILL; c.alignment = CENTER_ALIGN

    # Year index row (0 .. total_years)
    ws.cell(row=year_index_row, column=2, value="Year index (i)").font = ITALIC_FONT
    for i in range(total_years + 1):
        col = 3 + i
        c = ws.cell(row=year_index_row, column=col, value=i)
        c.font = ITALIC_FONT; c.number_format = "0"; c.alignment = RIGHT_ALIGN

    # Revenue growth row:
    #   Year 1  → consensus Y1 growth (G13) if present, else trailing CAGR (C11). Plus scenario adj.
    #   Year 2  → consensus Y2 growth (G14) if present, else trailing CAGR (C11). Plus scenario adj.
    #   Year 3+ → linear fade from Year-2 growth down to terminal growth (C21).
    # Scenario adj is picked via INDEX/MATCH on the scenario selector (C40).
    ws.cell(row=growth_row, column=2, value="Revenue growth").font = LABEL_FONT
    ws.cell(row=growth_row, column=3, value=dcf.trailing_revenue_cagr).number_format = FMT_PCT1  # Y0 actual
    ws.cell(row=growth_row, column=3).font = LABEL_FONT

    # G1 = IFERROR(G13, C11) + scen_adj   ;   G2 = IFERROR(G14, C11) + scen_adj
    g_y1_expr = f"(IFERROR($G$13,$C$11)+{scen_lookup(35)})"
    g_y2_expr = f"(IFERROR($G$14,$C$11)+{scen_lookup(35)})"

    for i in range(total_years):
        col = 4 + i
        year_count = i + 1            # 1-indexed year
        L = get_column_letter(col)
        if year_count == 1:
            formula = f"={g_y1_expr}"
        elif year_count == 2:
            formula = f"={g_y2_expr}"
        else:
            # Linear fade from Year-2 growth to terminal growth (C21) over remaining years
            # frac = (year_count - 2) / (total_years - 2)
            remaining = total_years - 2  # remaining years after Y2
            frac = (year_count - 2) / remaining if remaining > 0 else 1
            formula = f"={g_y2_expr}-({g_y2_expr}-$C$21)*{frac:.6f}"
        c = ws.cell(row=growth_row, column=col, value=formula)
        c.font = LABEL_FONT; c.number_format = FMT_PCT1; c.alignment = RIGHT_ALIGN

    # Revenue — Y0 = base revenue, Yn = prev × (1 + growth)
    ws.cell(row=revenue_row, column=2, value="Revenue").font = BOLD_LABEL_FONT
    c = ws.cell(row=revenue_row, column=3, value="=$C$31")
    c.font = TOTAL_FONT; c.number_format = FMT_NUM; c.alignment = RIGHT_ALIGN
    for i in range(total_years):
        col = 4 + i
        prev_col = get_column_letter(col - 1)
        formula = f"={prev_col}{revenue_row}*(1+{get_column_letter(col)}{growth_row})"
        c = ws.cell(row=revenue_row, column=col, value=formula)
        c.font = TOTAL_FONT; c.number_format = FMT_NUM; c.alignment = RIGHT_ALIGN

    # EBIT margin row — mirrors revenue-growth logic using consensus Y1/Y2:
    #   Year 1 → IFERROR(G15, (C12 + IFERROR(G16, C12))/2) + scen_margin_premium
    #            (G15 = Y1 target from EPS; if missing, bridge halfway between current and Y2)
    #   Year 2 → IFERROR(G16, C12) + scen_margin_premium   (G16 = Y2 target from EPS)
    #   Year 3+ → linear expansion from Y2 anchor to (Y2 anchor + 3%) over remaining years
    #             + scen_margin_premium
    ws.cell(row=margin_row, column=2, value="EBIT margin").font = LABEL_FONT
    c = ws.cell(row=margin_row, column=3, value="=$C$12")
    c.font = LABEL_FONT; c.number_format = FMT_PCT1; c.alignment = RIGHT_ALIGN

    scen_margin = scen_lookup(36)
    # Y1 margin expression: prefer G15; fall back to midpoint of C12 and G16; else C12
    m_y1_expr = f"(IFERROR($G$15,IFERROR(($C$12+$G$16)/2,$C$12))+{scen_margin})"
    # Y2 margin expression: prefer G16; fall back to C12
    m_y2_anchor_expr = f"IFERROR($G$16,$C$12)"
    m_y2_expr = f"({m_y2_anchor_expr}+{scen_margin})"
    # Terminal (year 10) margin: Y2 anchor + 300bps expansion + premium
    m_terminal_expr = f"({m_y2_anchor_expr}+0.03+{scen_margin})"

    for i in range(total_years):
        col = 4 + i
        year_count = i + 1
        if year_count == 1:
            formula = f"={m_y1_expr}"
        elif year_count == 2:
            formula = f"={m_y2_expr}"
        else:
            remaining = total_years - 2
            frac = (year_count - 2) / remaining if remaining > 0 else 1
            formula = f"={m_y2_expr}+({m_terminal_expr}-{m_y2_expr})*{frac:.6f}"
        c = ws.cell(row=margin_row, column=col, value=formula)
        c.font = LABEL_FONT; c.number_format = FMT_PCT1; c.alignment = RIGHT_ALIGN

    # EBIT = revenue × margin
    ws.cell(row=ebit_row, column=2, value="EBIT").font = LABEL_FONT
    for col in range(3, last_col_idx + 1):
        L = get_column_letter(col)
        formula = f"={L}{revenue_row}*{L}{margin_row}"
        c = ws.cell(row=ebit_row, column=col, value=formula)
        c.font = LABEL_FONT; c.number_format = FMT_NUM; c.alignment = RIGHT_ALIGN

    # Taxes = -EBIT × tax_rate   (shown negative)
    ws.cell(row=tax_row, column=2, value="(−) Taxes").font = LABEL_FONT
    for col in range(3, last_col_idx + 1):
        L = get_column_letter(col)
        formula = f"=-{L}{ebit_row}*$C$8"
        c = ws.cell(row=tax_row, column=col, value=formula)
        c.font = LABEL_FONT; c.number_format = FMT_NUM; c.alignment = RIGHT_ALIGN

    # NOPAT = EBIT + Tax
    ws.cell(row=nopat_row, column=2, value="NOPAT").font = TOTAL_FONT
    for col in range(3, last_col_idx + 1):
        L = get_column_letter(col)
        formula = f"={L}{ebit_row}+{L}{tax_row}"
        c = ws.cell(row=nopat_row, column=col, value=formula)
        c.font = TOTAL_FONT; c.number_format = FMT_NUM; c.alignment = RIGHT_ALIGN
    for cc in range(2, last_col_idx + 1):
        ws.cell(row=nopat_row, column=cc).border = TOP_BORDER

    # (+) D&A = revenue × D&A%
    ws.cell(row=da_row, column=2, value="(+) D&A").font = LABEL_FONT
    for col in range(3, last_col_idx + 1):
        L = get_column_letter(col)
        formula = f"={L}{revenue_row}*$C$13"
        c = ws.cell(row=da_row, column=col, value=formula)
        c.font = LABEL_FONT; c.number_format = FMT_NUM; c.alignment = RIGHT_ALIGN

    # (-) CapEx = -revenue × fade(trailing=C14, terminal=D14, year_count)
    # Fade logic: years 1-2 use trailing; years 3 to 2+fade_years interpolate;
    # year 2+fade_years+ uses terminal. If trailing ≤ terminal, no fade needed.
    # Fade is hardcoded to 5 years (per config default); user can override D14.
    ws.cell(row=capex_row, column=2, value="(−) CapEx (fade: trailing → terminal)").font = LABEL_FONT
    fade_years = dcf.assumptions.capex_fade_years
    for col in range(3, last_col_idx + 1):
        L = get_column_letter(col)
        year_count = col - 3   # col 3 = Y0, col 4 = Y1, etc.
        if year_count <= 2:
            # Trailing
            ratio_expr = "$C$14"
        elif year_count >= 2 + fade_years:
            # Terminal (but if trailing < terminal, stick with trailing)
            ratio_expr = "IF($C$14<=$D$14,$C$14,$D$14)"
        else:
            frac = (year_count - 2) / fade_years
            ratio_expr = (f"IF($C$14<=$D$14,$C$14,"
                          f"$C$14-($C$14-$D$14)*{frac:.6f})")
        formula = f"=-{L}{revenue_row}*{ratio_expr}"
        c = ws.cell(row=capex_row, column=col, value=formula)
        c.font = LABEL_FONT; c.number_format = FMT_NUM; c.alignment = RIGHT_ALIGN

    # (-) SBC = -revenue × fade(trailing=C15, terminal=D15)
    ws.cell(row=sbc_row, column=2,
            value="(−) Stock-based comp (fade: trailing → terminal)"
            if dcf.assumptions.sbc_as_cash_expense else "SBC (NOT deducted)").font = LABEL_FONT
    sbc_fade_years = dcf.assumptions.sbc_fade_years
    for col in range(3, last_col_idx + 1):
        L = get_column_letter(col)
        year_count = col - 3
        if not dcf.assumptions.sbc_as_cash_expense:
            c = ws.cell(row=sbc_row, column=col, value=0)
            c.font = LABEL_FONT; c.number_format = FMT_NUM; c.alignment = RIGHT_ALIGN
            continue
        if year_count <= 2:
            ratio_expr = "$C$15"
        elif year_count >= 2 + sbc_fade_years:
            ratio_expr = "IF($C$15<=$D$15,$C$15,$D$15)"
        else:
            frac = (year_count - 2) / sbc_fade_years
            ratio_expr = (f"IF($C$15<=$D$15,$C$15,"
                          f"$C$15-($C$15-$D$15)*{frac:.6f})")
        formula = f"=-{L}{revenue_row}*{ratio_expr}"
        c = ws.cell(row=sbc_row, column=col, value=formula)
        c.font = LABEL_FONT; c.number_format = FMT_NUM; c.alignment = RIGHT_ALIGN

    # (-) ΔNWC = -MAX(revenue - prev_revenue, 0) × NWC%
    ws.cell(row=nwc_row, column=2, value="(−) Δ NWC").font = LABEL_FONT
    c = ws.cell(row=nwc_row, column=3, value=0)
    c.font = LABEL_FONT; c.number_format = FMT_NUM; c.alignment = RIGHT_ALIGN
    for i in range(total_years):
        col = 4 + i
        L = get_column_letter(col)
        prev_L = get_column_letter(col - 1)
        formula = f"=-MAX({L}{revenue_row}-{prev_L}{revenue_row},0)*$C$16"
        c = ws.cell(row=nwc_row, column=col, value=formula)
        c.font = LABEL_FONT; c.number_format = FMT_NUM; c.alignment = RIGHT_ALIGN

    # FCFF = NOPAT + D&A + CapEx + SBC + NWC  (signs already baked in)
    ws.cell(row=fcff_row, column=2, value="Free Cash Flow to Firm (FCFF)").font = TOTAL_FONT
    for col in range(3, last_col_idx + 1):
        L = get_column_letter(col)
        formula = f"=SUM({L}{nopat_row}:{L}{nwc_row})"
        c = ws.cell(row=fcff_row, column=col, value=formula)
        c.font = TOTAL_FONT; c.number_format = FMT_NUM; c.alignment = RIGHT_ALIGN
    for cc in range(2, last_col_idx + 1):
        ws.cell(row=fcff_row, column=cc).border = TOP_BORDER

    # Discount factor = 1 / (1 + WACC)^i
    ws.cell(row=disc_row, column=2, value="Discount factor  1/(1+WACC)^i").font = LABEL_FONT
    for col in range(3, last_col_idx + 1):
        L = get_column_letter(col)
        formula = f"=1/(1+$C$49)^{L}{year_index_row}"
        c = ws.cell(row=disc_row, column=col, value=formula)
        c.font = LABEL_FONT; c.number_format = FMT_PCT1; c.alignment = RIGHT_ALIGN

    # PV of FCFF = FCFF × discount
    ws.cell(row=pv_row, column=2, value="PV of FCFF").font = TOTAL_FONT
    for col in range(3, last_col_idx + 1):
        L = get_column_letter(col)
        formula = f"={L}{fcff_row}*{L}{disc_row}"
        c = ws.cell(row=pv_row, column=col, value=formula)
        c.font = TOTAL_FONT; c.number_format = FMT_NUM; c.alignment = RIGHT_ALIGN

    # EBITDA = EBIT + D&A (for exit multiple)
    ws.cell(row=ebitda_row, column=2, value="EBITDA  (=EBIT + D&A)").font = ITALIC_FONT
    for col in range(3, last_col_idx + 1):
        L = get_column_letter(col)
        formula = f"={L}{ebit_row}+{L}{da_row}"
        c = ws.cell(row=ebitda_row, column=col, value=formula)
        c.font = ITALIC_FONT; c.number_format = FMT_NUM; c.alignment = RIGHT_ALIGN

    # ============================================================
    # BLOCK 8 — Terminal Value (rows 71-78)
    # ============================================================
    _section(71, "Terminal Value — Dual Method")
    last_L = get_column_letter(last_col_idx)
    _calc(72, f"Sum of PV of FCFF (Y1–Y{total_years})",
          f"=SUM(D{pv_row}:{last_L}{pv_row})", FMT_NUM)
    _calc(73, f"Year-{total_years} FCFF",
          f"={last_L}{fcff_row}", FMT_NUM)
    _calc(74, f"Year-{total_years} EBITDA",
          f"={last_L}{ebitda_row}", FMT_NUM)
    _calc(75, "Terminal Value — Gordon Growth",
          f"=C73*(1+$C$21)/($C$49-$C$21)", FMT_NUM)
    _calc(76, "PV of Gordon Growth TV",
          f"=C75*{last_L}{disc_row}", FMT_NUM)
    _calc(77, "Terminal Value — Exit Multiple  (Y10 EBITDA × peer EV/EBITDA)",
          f"=C74*$C$24", FMT_NUM)
    _calc(78, "PV of Exit Multiple TV",
          f"=C77*{last_L}{disc_row}", FMT_NUM)
    # When Exit Multiple TV is zero (no peer EV/EBITDA data), fall back to
    # 100% Gordon Growth instead of penalizing with a 50/50 blend against zero.
    _calc(79, "Blended PV of Terminal Value  (weighted)",
          f"=IF(C78=0,C76,C76*$C$22+C78*$C$23)", FMT_NUM, bold=True)
    for cc in range(2, 4):
        ws.cell(row=79, column=cc).border = TOP_BORDER

    # ============================================================
    # BLOCK 9 — Equity Bridge (rows 82-89)
    # ============================================================
    _section(82, "Equity Bridge")
    _calc(83, "Sum of PV of FCFF ($M)",     "=C72", FMT_NUM)
    _calc(84, "+ PV of Terminal Value ($M)", "=C79", FMT_NUM)
    _calc(85, "Enterprise Value ($M)",       "=C83+C84", FMT_NUM, bold=True)
    for cc in range(2, 4):
        ws.cell(row=85, column=cc).border = TOP_BORDER
    _calc(86, "+ Cash and equivalents",      "=C29", FMT_NUM)
    _calc(87, "− Total debt",                 "=-C28", FMT_NUM)
    _calc(88, "Implied Equity Value ($M)",   "=C85+C86+C87", FMT_NUM, bold=True)
    for cc in range(2, 4):
        ws.cell(row=88, column=cc).border = TOP_BORDER
    _calc(89, "Shares outstanding (mm)",     "=C30", FMT_SHARES)
    _calc(90, "Implied Price per Share",     "=C88/C89", FMT_USD, bold=True)
    for cc in range(2, 4):
        ws.cell(row=90, column=cc).border = DOUBLE_TOP

    if current_share_price:
        _input(92, "Current market price", current_share_price, FMT_USD)
        _calc(93, "Upside / (Downside)", "=C90/C92-1", FMT_PCT1, bold=True)

    # ============================================================
    # BLOCK 10 — Scenario Comparison (rows 96-102) — values from Python
    # (explanation cell says: change C40 selector to see each live)
    # ============================================================
    _section(96, "Scenario Comparison (pre-computed at run time — change C40 to switch live model)")
    scen_header = ["", "Bear", "Base", "Bull"]
    for i, h in enumerate(scen_header):
        c = ws.cell(row=97, column=2 + i, value=h)
        c.font = HEADER_FONT; c.fill = HEADER_FILL; c.alignment = CENTER_ALIGN
    scen_order = ["Bear", "Base", "Bull"]
    scen_rows_out = [
        ("WACC",            lambda s: s.wacc,                              FMT_PCT1),
        ("EV ($M)",         lambda s: s.enterprise_value / 1e6,             FMT_NUM),
        ("Equity ($M)",     lambda s: s.equity_value / 1e6,                 FMT_NUM),
        ("Price / share",   lambda s: s.price_per_share,                    FMT_USD),
        ("Upside",          lambda s: (s.price_per_share / current_share_price - 1)
                                    if current_share_price else None,       FMT_PCT1),
    ]
    for j, (label, fn, fmt) in enumerate(scen_rows_out, start=98):
        ws.cell(row=j, column=2, value=label).font = LABEL_FONT
        for i, name in enumerate(scen_order):
            v = fn(dcf.scenarios[name])
            if v is None:
                continue
            c = ws.cell(row=j, column=3 + i, value=v)
            c.number_format = fmt
            c.font = TOTAL_FONT if label == "Price / share" else LABEL_FONT
            c.alignment = RIGHT_ALIGN

    # ============================================================
    # BLOCK 11 — Sensitivity tables
    # Placed AFTER the Football Field (which lives at rows 105-130) so users
    # see the football field visually before scrolling to sensitivities.
    # ============================================================
    sens_col = 2   # column B (left-aligned)
    sens_row = 135
    _section(sens_row, "Sensitivity 1: Price/Share — WACC × Terminal Growth", col=sens_col)
    wacc_grid = sorted(set(k[0] for k in dcf.sensitivity_wacc_g.keys()))
    g_grid = sorted(set(k[1] for k in dcf.sensitivity_wacc_g.keys()))
    for i, g in enumerate(g_grid, start=sens_col + 1):
        c = ws.cell(row=sens_row + 1, column=i, value=g)
        c.font = BOLD_LABEL_FONT; c.number_format = FMT_PCT1; c.alignment = CENTER_ALIGN
    for j, wacc in enumerate(wacc_grid, start=sens_row + 2):
        c = ws.cell(row=j, column=sens_col, value=wacc)
        c.font = BOLD_LABEL_FONT; c.number_format = FMT_PCT1
        for i, g in enumerate(g_grid, start=sens_col + 1):
            val = dcf.sensitivity_wacc_g.get((round(wacc, 4), round(g, 4)))
            if val is not None:
                sc = ws.cell(row=j, column=i, value=val)
                sc.number_format = FMT_USD; sc.font = LABEL_FONT; sc.alignment = RIGHT_ALIGN

    sens2_row = sens_row + len(wacc_grid) + 4
    _section(sens2_row, "Sensitivity 2: Price/Share — WACC × Exit Multiple (EV/EBITDA)", col=sens_col)
    if dcf.sensitivity_wacc_exit:
        wacc_grid2 = sorted(set(k[0] for k in dcf.sensitivity_wacc_exit.keys()))
        em_grid = sorted(set(k[1] for k in dcf.sensitivity_wacc_exit.keys()))
        for i, em in enumerate(em_grid, start=sens_col + 1):
            c = ws.cell(row=sens2_row + 1, column=i, value=em)
            c.font = BOLD_LABEL_FONT; c.number_format = "0.0\"x\""; c.alignment = CENTER_ALIGN
        for j, wacc in enumerate(wacc_grid2, start=sens2_row + 2):
            c = ws.cell(row=j, column=sens_col, value=wacc)
            c.font = BOLD_LABEL_FONT; c.number_format = FMT_PCT1
            for i, em in enumerate(em_grid, start=sens_col + 1):
                val = dcf.sensitivity_wacc_exit.get((round(wacc, 4), round(em, 4)))
                if val is not None:
                    sc = ws.cell(row=j, column=i, value=val)
                    sc.number_format = FMT_USD; sc.font = LABEL_FONT; sc.alignment = RIGHT_ALIGN
    else:
        # No peer EV/EBITDA data available — show explanatory note instead of empty grid.
        note_row = sens2_row + 1
        note_text = ("Not shown — no peer with valid TTM EV/EBITDA. "
                     "All selected peers are pre-profit (negative or zero EBITDA), "
                     "so no reliable peer-based exit multiple can be anchored. "
                     "Rely on Sensitivity 1 (Gordon Growth) and the football-field "
                     "analyst target range above for terminal-value cross-checks.")
        c = ws.cell(row=note_row, column=sens_col, value=note_text)
        c.font = ITALIC_FONT
        c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        # Merge across columns B-H for readability
        ws.merge_cells(start_row=note_row, start_column=sens_col,
                       end_row=note_row + 2, end_column=sens_col + 6)
        ws.row_dimensions[note_row].height = 45

    # ============================================================
    # BLOCK 12 — Football Field chart (embedded PNG)
    # Positioned just below the Scenario Comparison block (row 96-102) so the
    # user sees it before the sensitivity tables. Occupies rows 105-132.
    # ============================================================
    if football_field_png and football_field_png.exists():
        img_row = 105
        _section(img_row, "Football Field — Valuation Summary")
        from openpyxl.drawing.image import Image as XLImage
        img = XLImage(str(football_field_png))
        img.width = 720
        img.height = 360
        ws.add_image(img, f"B{img_row + 1}")


# ---------- Public entry ----------

def build_workbook(stmt: IncomeStatement, dcf: "DCFResult", output_path: Path,
                   current_share_price: Optional[float] = None,
                   current_market_cap: Optional[float] = None,
                   football_field_png: Optional[Path] = None) -> Path:
    wb = Workbook()
    ws_is = wb.active
    ws_is.title = stmt.ticker[:31]

    _write_is_sheet(ws_is, stmt, dcf)

    ws_val = wb.create_sheet("Valuation")
    _write_valuation_sheet(ws_val, stmt, dcf, current_share_price, current_market_cap,
                           football_field_png=football_field_png)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return output_path


if __name__ == "__main__":
    pass
