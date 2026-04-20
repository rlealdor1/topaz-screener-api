"""Build the Comps workbook — peers table with TTM fundamentals and multiples."""
from __future__ import annotations
from pathlib import Path
from statistics import mean, median
from typing import List

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from ..data.yfinance_client import QuoteSnapshot


HEADER_FILL = PatternFill("solid", fgColor="1F3864")
HEADER_FONT = Font(name="Calibri", size=10, color="FFFFFF", bold=True)
SUBJECT_FILL = PatternFill("solid", fgColor="FFF2CC")  # light yellow for target company
MED_FILL = PatternFill("solid", fgColor="F4B084")       # peach for median/mean rows
LABEL_FONT = Font(name="Calibri", size=10)
BOLD_FONT = Font(name="Calibri", size=10, bold=True)

FMT_NUM = "#,##0;(#,##0);\"-\""
FMT_MULT = "0.0\"x\""
FMT_PCT = "0%"
FMT_USD = "$#,##0.00"

COLUMNS = [
    ("Company Name",        lambda q: q.company_name or q.ticker,              None,     35),
    ("Sector / Industry",   lambda q: (q.industry or q.sector or "-")[:35],    None,     28),
    ("HQ",                  lambda q: (q.hq_country or "-")[:15],              None,     12),
    ("Current Share Price", lambda q: q.current_price,                         FMT_USD,  15),
    ("52-Week High",        lambda q: q.week52_high,                           FMT_USD,  13),
    ("% off High",          lambda q: q.pct_off_52w_high,                      FMT_PCT,  11),
    ("Market Cap ($M)",     lambda q: _to_m(q.market_cap),                     FMT_NUM,  14),
    ("EV ($M)",             lambda q: _to_m(q.enterprise_value),               FMT_NUM,  13),
    ("Revenue ($M)",        lambda q: _to_m(q.revenue_ttm),                    FMT_NUM,  14),
    ("Gross Profit ($M)",   lambda q: _to_m(q.gross_profit_ttm),               FMT_NUM,  15),
    ("EBITDA ($M)",         lambda q: _to_m(q.ebitda_ttm),                     FMT_NUM,  13),
    ("Cash ($M)",           lambda q: _to_m(q.cash),                           FMT_NUM,  12),
    ("Total Debt ($M)",     lambda q: _to_m(q.total_debt),                     FMT_NUM,  14),
    ("FCF ($M)",            lambda q: _to_m(q.free_cash_flow),                 FMT_NUM,  12),
    ("CapEx ($M)",          lambda q: _to_m(q.capex),                          FMT_NUM,  12),
    ("Operating CF ($M)",   lambda q: _to_m(q.operating_cash_flow),            FMT_NUM,  15),
    ("Net Leverage",        lambda q: q.net_leverage,                          FMT_MULT, 12),
    ("P/E",                 lambda q: q.pe_ratio,                              FMT_MULT, 10),
    ("EV / EBITDA",         lambda q: q.ev_ebitda,                             FMT_MULT, 12),
    ("EV / Sales",          lambda q: q.ev_sales,                              FMT_MULT, 11),
]


def _to_m(v):
    if v is None:
        return None
    return v / 1_000_000


def _set_header(ws, row: int):
    for i, (name, _, _, width) in enumerate(COLUMNS, start=1):
        c = ws.cell(row=row, column=i, value=name)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = Alignment(horizontal="center", wrap_text=True, vertical="center")
        ws.column_dimensions[get_column_letter(i)].width = width
    ws.row_dimensions[row].height = 30


def _write_row(ws, row_idx: int, q: QuoteSnapshot, fill=None):
    for i, (name, extractor, fmt, _) in enumerate(COLUMNS, start=1):
        v = extractor(q)
        c = ws.cell(row=row_idx, column=i, value=v)
        if fmt:
            c.number_format = fmt
        c.font = LABEL_FONT
        c.alignment = Alignment(horizontal="right" if isinstance(v, (int, float)) else "left")
        if fill:
            c.fill = fill


def _write_stat_row(ws, row_idx: int, label: str, quotes: List[QuoteSnapshot], fn):
    """Write Median/Mean row across all peer quotes."""
    c = ws.cell(row=row_idx, column=1, value=label)
    c.font = BOLD_FONT
    c.fill = MED_FILL
    c.alignment = Alignment(horizontal="left")
    for i, (name, extractor, fmt, _) in enumerate(COLUMNS, start=1):
        if i <= 3:
            ws.cell(row=row_idx, column=i).fill = MED_FILL
            continue
        values = [extractor(q) for q in quotes]
        values = [v for v in values if isinstance(v, (int, float))]
        if values:
            try:
                v = fn(values)
            except Exception:
                v = None
        else:
            v = None
        cell = ws.cell(row=row_idx, column=i, value=v)
        if fmt:
            cell.number_format = fmt
        cell.font = BOLD_FONT
        cell.fill = MED_FILL
        cell.alignment = Alignment(horizontal="right")


def build_comps_workbook(subject: QuoteSnapshot, peers: List[QuoteSnapshot],
                         output_path: Path) -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = "Comparables"

    # Row 1: title
    title_cell = ws.cell(row=1, column=1, value=f"{subject.ticker} - Comparables (TTM, USD)")
    title_cell.font = Font(name="Calibri", size=12, bold=True, color="1F3864")

    header_row = 3
    _set_header(ws, header_row)

    row = header_row + 1
    _write_row(ws, row, subject, fill=SUBJECT_FILL)
    row += 1
    for p in peers:
        _write_row(ws, row, p)
        row += 1

    # Median + Mean rows (across peers only)
    _write_stat_row(ws, row, "Median", peers, median)
    row += 1
    _write_stat_row(ws, row, "Mean", peers, mean)

    # Freeze header
    ws.freeze_panes = f"A{header_row + 1}"

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return output_path


if __name__ == "__main__":
    from src.data.yfinance_client import get_quote
    subject = get_quote("COIN")
    peer_tickers = ["HOOD", "CME", "NDAQ", "ICE", "SQ", "MARA", "RIOT"]
    peers = [get_quote(t) for t in peer_tickers]
    out = build_comps_workbook(subject, peers, Path("Auto-generated/COIN/COIN Comps.xlsx"))
    print(f"Wrote: {out.resolve()}")
